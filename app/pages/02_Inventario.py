# =========================================================
# IMPORTACIONES
# =========================================================

import sys
from io import BytesIO
from pathlib import Path

import pandas as pd
import plotly.express as px
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
    sys.path.append(str(ROOT))

DATA_DIR = ROOT / "data" / "processed"

from domain.inventory.metrics import (
    antiguedad_equipos,
    calcular_kpis,
    equipos_por_marca,
    equipos_por_tipo,
    obtener_alertas,
)
from domain.inventory.transform import aplicar_filtros
from services.inventory_service import obtener_inventario


def aplicar_estilo_excel(workbook):
    header_fill = PatternFill(fill_type="solid", fgColor="E8EEF5")
    alternate_fill = PatternFill(fill_type="solid", fgColor="F7FAFC")
    header_font = Font(name="Calibri", size=10, bold=True, color="23364D")
    body_font = Font(name="Calibri", size=10, color="23364D")
    border = Border(
        left=Side(style="thin", color="D0D7DE"),
        right=Side(style="thin", color="D0D7DE"),
        top=Side(style="thin", color="D0D7DE"),
        bottom=Side(style="thin", color="D0D7DE"),
    )
    align_center = Alignment(horizontal="center", vertical="center")

    for ws in workbook.worksheets:
        ws.sheet_view.zoomScale = 90
        ws.auto_filter.ref = ws.dimensions

        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = align_center
                else:
                    cell.font = body_font
                    if cell.row % 2 == 0:
                        cell.fill = alternate_fill

        for col_idx, col in enumerate(ws.columns, start=1):
            values = [cell.value for cell in col if cell.value is not None]
            if values:
                max_length = max(len(str(v)) for v in values)
            else:
                max_length = 10
            ws.column_dimensions[get_column_letter(col_idx)].width = min(
                max(12, max_length + 2), 50
            )


def exportar_excel(
    df_filtrado,
    alertas,
):

    output = BytesIO()

    with pd.ExcelWriter(
        output,
        engine="openpyxl",
    ) as writer:
        resumen = pd.DataFrame(
            {
                "Indicador": [
                    "Total Equipos",
                    "Sin Factura",
                    "Sin Fecha",
                    "Sin Serie",
                    "Sin Asignar",
                ],
                "Cantidad": [
                    len(df_filtrado),
                    len(alertas["sin_factura"]),
                    len(alertas["sin_fecha_compra"]),
                    len(alertas["sin_serie"]),
                    len(alertas["sin_asignar"]),
                ],
            }
        )

        resumen.to_excel(
            writer,
            sheet_name="Resumen",
            index=False,
        )

        df_filtrado.to_excel(
            writer,
            sheet_name="Inventario",
            index=False,
        )

        alertas["sin_factura"].to_excel(
            writer,
            sheet_name="Sin_Factura",
            index=False,
        )

        alertas["sin_fecha_compra"].to_excel(
            writer,
            sheet_name="Sin_Fecha",
            index=False,
        )

        alertas["sin_serie"].to_excel(
            writer,
            sheet_name="Sin_Serie",
            index=False,
        )

        alertas["sin_asignar"].to_excel(
            writer,
            sheet_name="Sin_Asignar",
            index=False,
        )

        workbook = writer.book
        aplicar_estilo_excel(workbook)

    output.seek(0)

    return output


# =========================================================
# CONFIGURACIÓN PÁGINA
# =========================================================

st.set_page_config(
    page_title="Inventario",
    page_icon="💻",
    layout="wide",
)

st.title("💻 Inventario de Equipos")

# =========================================================
# CARGAR DATOS
# =========================================================

ruta_datos = DATA_DIR / "inventario.csv"

if not ruta_datos.exists():
    st.warning(
        "No existen registros en el inventario. Usa el botón de actualizar para sincronizar desde Notion."
    )
    st.stop()

df = pd.read_csv(ruta_datos)

if df.empty:
    st.warning("No existen registros en el inventario.")
    st.stop()

# =========================================================
# SIDEBAR
# =========================================================

st.sidebar.header("Inventario")

if st.sidebar.button(
    "🔄 Actualizar Inventario",
    width="stretch",
):
    with st.spinner("Actualizando inventario desde Notion..."):
        obtener_inventario()

    st.success("Inventario actualizado correctamente.")

    st.rerun()

st.sidebar.divider()

# =========================================================
# FILTROS
# =========================================================

st.sidebar.subheader("Filtros")

tipos = sorted(df["Tipo"].dropna().unique())

tipo_seleccionado = st.sidebar.selectbox(
    "Tipo",
    ["Todos"] + list(tipos),
)

marcas = sorted(df["Marca"].dropna().unique())

marca_seleccionada = st.sidebar.selectbox(
    "Marca",
    ["Todas"] + list(marcas),
)

# =========================================================
# APLICAR FILTROS
# =========================================================

df_filtrado = aplicar_filtros(
    df=df,
    tipo=tipo_seleccionado,
    marca=marca_seleccionada,
)
# =========================================================
# ALERTAS Y EXPORTACIÓN
# =========================================================

alertas = obtener_alertas(df_filtrado)

excel_file = exportar_excel(
    df_filtrado,
    alertas,
)

# =========================================================
# KPIs
# =========================================================

kpis = calcular_kpis(df_filtrado)

col1, col2, col3, col4, col5 = st.columns(5)

with col1:
    st.metric(
        "Equipos",
        kpis["total_equipos"],
    )

with col2:
    st.metric(
        "Asignados",
        kpis["equipos_asignados"],
    )

with col3:
    st.metric(
        "Disponibles",
        kpis["sin_asignar"],
    )

with col4:
    st.metric(
        "Marcas",
        kpis["total_marcas"],
    )

with col5:
    st.metric(
        "Sin Serie",
        kpis["sin_serie"],
    )

st.divider()


# =========================================================
# GRÁFICOS
# =========================================================

col_graf1, col_graf2 = st.columns(2)

# ---------------------------------------------------------
# EQUIPOS POR TIPO
# ---------------------------------------------------------

with col_graf1:
    st.subheader("📊 Equipos por Tipo")

    tipo_chart = equipos_por_tipo(df_filtrado).reset_index()

    tipo_chart.columns = [
        "Tipo",
        "Cantidad",
    ]

    fig_tipo = px.bar(
        tipo_chart,
        x="Cantidad",
        y="Tipo",
        text="Cantidad",
        orientation="h",
        title="Distribución por Tipo",
    )

    fig_tipo.update_layout(
        height=350,
        xaxis_title="Cantidad",
        yaxis_title="",
        showlegend=False,
    )

    fig_tipo.update_traces(
        textposition="auto",
    )

    st.plotly_chart(
        fig_tipo,
        width="stretch",
    )

# ---------------------------------------------------------
# EQUIPOS POR MARCA
# ---------------------------------------------------------

with col_graf2:
    st.subheader("🏷️ Equipos por Marca")

    marca_chart = equipos_por_marca(df_filtrado).reset_index()

    marca_chart.columns = [
        "Marca",
        "Cantidad",
    ]

    fig_marca = px.bar(
        marca_chart,
        x="Cantidad",
        y="Marca",
        text="Cantidad",
        orientation="h",
        title="Distribución por Marca",
    )

    fig_marca.update_layout(
        height=350,
        xaxis_title="Cantidad",
        yaxis_title="",
        showlegend=False,
    )

    fig_marca.update_traces(
        textposition="auto",
    )

    st.plotly_chart(
        fig_marca,
        width="stretch",
    )

# =========================================================
# ANTIGÜEDAD DE EQUIPOS
# =========================================================

st.subheader("📅 Antigüedad de Equipos")

antiguedad_chart = antiguedad_equipos(df_filtrado).reset_index()

antiguedad_chart.columns = [
    "Rango",
    "Cantidad",
]

fig_antiguedad = px.bar(
    antiguedad_chart,
    x="Cantidad",
    y="Rango",
    text="Cantidad",
    orientation="h",
    title="Distribución por Antigüedad",
)

fig_antiguedad.update_layout(
    height=350,
    xaxis_title="Antigüedad",
    yaxis_title="Cantidad",
    showlegend=False,
)

fig_antiguedad.update_traces(
    textposition="outside",
)

st.plotly_chart(
    fig_antiguedad,
    width="stretch",
)

st.divider()

# =========================================================
# BOTON DESCARGAR EXCEL
# =========================================================

st.download_button(
    label="📥 Descargar Reporte Excel",
    data=excel_file,
    file_name="inventario_reporte.xlsx",
    mime=("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    width="stretch",
)

st.divider()

# =========================================================
# TABLA PRINCIPAL
# =========================================================

st.subheader(f"Inventario ({len(df_filtrado)} registros)")

st.dataframe(
    df_filtrado,
    width="stretch",
)

# =========================================================
# CALIDAD DE DATOS
# =========================================================

st.divider()

st.subheader("🚨 Calidad de Datos")

col_a1, col_a2, col_a3, col_a4 = st.columns(4)

with col_a1:
    st.metric(
        "Sin Serie",
        len(alertas["sin_serie"]),
    )

with col_a2:
    st.metric(
        "Sin Factura",
        len(alertas["sin_factura"]),
    )

with col_a3:
    st.metric(
        "Sin Fecha",
        len(alertas["sin_fecha_compra"]),
    )

with col_a4:
    st.metric(
        "Sin Asignar",
        len(alertas["sin_asignar"]),
    )

st.divider()

columnas_alerta = [
    "Tipo",
    "Marca",
    "Modelo",
    "Serie",
    "Asignado",
]

alertas_config = {
    "sin_factura": "sin factura registrada",
    "sin_fecha_compra": "sin fecha de compra",
    "sin_serie": "sin número de serie",
    "sin_asignar": "sin asignar",
}

for clave, descripcion in alertas_config.items():
    if len(alertas[clave]) > 0:
        st.warning(f"⚠️ Existen {len(alertas[clave])} equipos {descripcion}.")

        columnas_disponibles = [
            col for col in columnas_alerta if col in alertas[clave].columns
        ]

        st.dataframe(
            alertas[clave][columnas_disponibles],
            width="stretch",
        )

if (
    len(alertas["sin_factura"]) == 0
    and len(alertas["sin_fecha_compra"]) == 0
    and len(alertas["sin_serie"]) == 0
    and len(alertas["sin_asignar"]) == 0
):
    st.success("✅ No se encontraron observaciones en el inventario.")

# ---------------------------------------------------------
# EQUIPOS SIN FACTURA
# ---------------------------------------------------------

if len(alertas["sin_factura"]) > 0:
    st.warning(
        f"⚠️ Existen {len(alertas['sin_factura'])} equipos sin factura registrada."
    )

    st.dataframe(
        alertas["sin_factura"][columnas_alerta],
        width="stretch",
    )

# ---------------------------------------------------------
# EQUIPOS SIN FECHA DE COMPRA
# ---------------------------------------------------------

if len(alertas["sin_fecha_compra"]) > 0:
    st.warning(
        f"⚠️ Existen {len(alertas['sin_fecha_compra'])} equipos sin fecha de compra."
    )

    st.dataframe(
        alertas["sin_fecha_compra"][columnas_alerta],
        width="stretch",
    )

# ---------------------------------------------------------
# EQUIPOS SIN SERIE
# ---------------------------------------------------------

if len(alertas["sin_serie"]) > 0:
    st.warning(f"⚠️ Existen {len(alertas['sin_serie'])} equipos sin número de serie.")

    st.dataframe(
        alertas["sin_serie"][columnas_alerta],
        width="stretch",
    )

# ---------------------------------------------------------
# EQUIPOS SIN ASIGNAR
# ---------------------------------------------------------

if len(alertas["sin_asignar"]) > 0:
    st.warning(f"⚠️ Existen {len(alertas['sin_asignar'])} equipos sin asignar.")

    st.dataframe(
        alertas["sin_asignar"][columnas_alerta],
        width="stretch",
    )

# ---------------------------------------------------------
# SIN OBSERVACIONES
# ---------------------------------------------------------

if (
    len(alertas["sin_factura"]) == 0
    and len(alertas["sin_fecha_compra"]) == 0
    and len(alertas["sin_serie"]) == 0
    and len(alertas["sin_asignar"]) == 0
):
    st.success("✅ No se encontraron observaciones en el inventario.")
