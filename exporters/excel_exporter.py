from pathlib import Path
from datetime import datetime

import pandas as pd

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule


# =========================================================
# CONFIGURACIÓN GENERAL
# =========================================================
FORMATO_MONEDA = '[$S/] #,##0.00'
FORMATO_FECHA = 'dd/mm/yyyy'
FORMATO_PORCENTAJE = '0.00%'

COLUMNAS_OCULTAR = ['categoria_id']
COLUMNAS_MONETARIAS = {
    'monto',
    'Ingresos',
    'Gastos',
    'Balance',
    'Valor',
    'Promedio mensual',
}

COLUMNAS_PORCENTAJE = {
    'Pct_Gastos',
    'Pct_Ahorro',
    'Variacion_Balance',
}


# =========================================================
# UTILIDADES
# =========================================================
def calcular_score_financiero(ingresos, gastos, balance):
    score = 100

    if balance < 0:
        score -= 40

    if ingresos > 0 and abs(gastos) > ingresos:
        score -= 30

    ratio_gasto = abs(gastos) / ingresos if ingresos > 0 else 1

    if ratio_gasto > 0.8:
        score -= 20

    return max(score, 0)


# =========================================================
# FORMATO DE HOJAS
# =========================================================
def crear_estilos():
    fill_header = PatternFill('solid', fgColor='1F4E78')  # azul oscuro
    font_header = Font(name='Calibri', size=10, color='FFFFFF', bold=True)
    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    border_thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )

    return {
        'fill_header': fill_header,
        'font_header': font_header,
        'align_center': align_center,
        'align_right': align_right,
        'border_thin': border_thin,
    }


def formatear_hoja(ws, estilos):
    fill_header = estilos['fill_header']
    font_header = estilos['font_header']
    align_center = estilos['align_center']
    align_right = estilos['align_right']
    border_thin = estilos['border_thin']

    # Configuración general
    ws.sheet_view.zoomScale = 80
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    # Fuente y bordes
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(
                name='Calibri',
                size=10,
                bold=cell.font.bold,
                italic=cell.font.italic,
                color=cell.font.color,
            )
            cell.border = border_thin

    # Encabezado
    for cell in ws[1]:
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_center
        cell.border = border_thin

    # Formatos y ancho automático
    for col_idx, col in enumerate(ws.columns, start=1):
        header = ws.cell(row=1, column=col_idx).value
        max_length = 0

        for cell in col:
            # Formato monetario
            if (
                cell.row > 1
                and isinstance(cell.value, (int, float))
                and header in COLUMNAS_MONETARIAS
            ):
                cell.number_format = FORMATO_MONEDA
                cell.alignment = align_right

            # Formato porcentual
            if (
                cell.row > 1
                and isinstance(cell.value, (int, float))
                and header in COLUMNAS_PORCENTAJE
            ):
                cell.number_format = FORMATO_PORCENTAJE
                cell.alignment = align_right

            # Formato fecha
            if cell.row > 1 and header == 'fecha' and cell.value is not None:
                cell.number_format = FORMATO_FECHA

            # Cálculo de ancho
            try:
                if cell.value is not None:
                    longitud = len(str(cell.value))
                    max_length = max(max_length, longitud)
            except Exception:
                pass

        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(12, min(max_length + 2, 50))


# =========================================================
# FORMATO CONDICIONAL
# =========================================================
def aplicar_formato_condicional_resumen_mensual(ws):
    col_map = {cell.value: cell.column for cell in ws[1]}
    fill_rojo = PatternFill('solid', fgColor='FFC7CE')

    # Balance negativo
    if 'Balance' in col_map:
        col = get_column_letter(col_map['Balance'])
        ws.conditional_formatting.add(
            f'{col}2:{col}{ws.max_row}',
            CellIsRule(operator='lessThan', formula=['0'], fill=fill_rojo),
        )

    # % Gastos > 80%
    if 'Pct_Gastos' in col_map:
        col = get_column_letter(col_map['Pct_Gastos'])
        ws.conditional_formatting.add(
            f'{col}2:{col}{ws.max_row}',
            CellIsRule(operator='greaterThan', formula=['0.8'], fill=fill_rojo),
        )

    # % Ahorro < 0
    if 'Pct_Ahorro' in col_map:
        col = get_column_letter(col_map['Pct_Ahorro'])
        ws.conditional_formatting.add(
            f'{col}2:{col}{ws.max_row}',
            CellIsRule(operator='lessThan', formula=['0'], fill=fill_rojo),
        )


# =========================================================
# EXPORTACIÓN PRINCIPAL
# =========================================================
def exportar_finanzas_excel(df, output_path=None):
    Path('dist').mkdir(exist_ok=True)

    if output_path is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = Path('dist') / f'finanzas_{timestamp}.xlsx'
    else:
        output_path = Path(output_path)

    # Copia y limpieza
    df = df.copy()
    df = df.drop(
        columns=[c for c in COLUMNAS_OCULTAR if c in df.columns],
        errors='ignore',
    )

    if 'fecha' in df.columns:
        df['fecha'] = pd.to_datetime(df['fecha'], errors='coerce')

    # KPIs
    ingresos = df.loc[df['monto'] > 0, 'monto'].sum() if 'monto' in df.columns else 0
    gastos = df.loc[df['monto'] < 0, 'monto'].sum() if 'monto' in df.columns else 0
    balance = ingresos + gastos
    total_movimientos = len(df)

    categoria_top = 'N/A'
    if {'categoria', 'monto'}.issubset(df.columns):
        gastos_df = df[df['monto'] < 0]
        if not gastos_df.empty:
            try:
                categoria_top = (
                    gastos_df.groupby('categoria')['monto']
                    .sum()
                    .idxmin()
                )
            except Exception:
                categoria_top = 'N/A'

    promedio_mensual = 0
    if {'mes', 'monto'}.issubset(df.columns):
        try:
            promedio_mensual = df.groupby('mes')['monto'].sum().mean()
        except Exception:
            promedio_mensual = 0

    score = calcular_score_financiero(ingresos, gastos, balance)

    # Hoja resumen
    resumen = pd.DataFrame(
        {
            'Indicador': [
                'Total de ingresos',
                'Total de gastos',
                'Balance neto',
                'Número de movimientos',
                'Categoría de mayor gasto',
                'Promedio mensual',
                'Score financiero',
            ],
            'Valor': [
                ingresos,
                gastos,
                balance,
                total_movimientos,
                categoria_top,
                promedio_mensual,
                score,
            ],
        }
    )

    estilos = crear_estilos()

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # -------------------------------------------------
        # Hoja Movimientos
        # -------------------------------------------------
        df.to_excel(writer, sheet_name='Movimientos', index=False)

        # -------------------------------------------------
        # Hoja Resumen
        # -------------------------------------------------
        resumen.to_excel(writer, sheet_name='Resumen', index=False)

        # -------------------------------------------------
        # Hojas analíticas
        # -------------------------------------------------
        if {'mes', 'categoria', 'monto'}.issubset(df.columns):
            gastos_df = df[df['monto'] < 0].copy()

            if not gastos_df.empty:
                # Pivot_Categorias
                pivot_categorias = pd.pivot_table(
                    gastos_df,
                    index='mes',
                    columns='categoria',
                    values='monto',
                    aggfunc='sum',
                    fill_value=0,
                )
                pivot_categorias.to_excel(writer, sheet_name='Pivot_Categorias')

                # Resumen_Mensual
                ingresos_m = df[df['monto'] > 0].groupby('mes')['monto'].sum()
                gastos_m = df[df['monto'] < 0].groupby('mes')['monto'].sum()

                resumen_mensual = pd.DataFrame(
                    {
                        'Ingresos': ingresos_m,
                        'Gastos': gastos_m,
                    }
                ).fillna(0)

                resumen_mensual['Balance'] = (
                    resumen_mensual['Ingresos'] + resumen_mensual['Gastos']
                )

                resumen_mensual['Pct_Gastos'] = resumen_mensual.apply(
                    lambda r: abs(r['Gastos']) / r['Ingresos'] if r['Ingresos'] != 0 else 0,
                    axis=1,
                )

                resumen_mensual['Pct_Ahorro'] = resumen_mensual.apply(
                    lambda r: r['Balance'] / r['Ingresos'] if r['Ingresos'] != 0 else 0,
                    axis=1,
                )

                resumen_mensual['Variacion_Balance'] = (
                    resumen_mensual['Balance'].pct_change().fillna(0)
                )

                resumen_mensual.to_excel(writer, sheet_name='Resumen_Mensual')

                # Top_Gastos
                top_gastos = gastos_df.sort_values('monto').head(50)
                top_gastos.to_excel(writer, sheet_name='Top_Gastos', index=False)
        # -------------------------------------------------
        # Hoja Dashboard Ejecutivo
        # -------------------------------------------------
        dashboard = pd.DataFrame({
            "Indicador": [
                "Total Ingresos",
                "Total Gastos",
                "Balance Neto",
                "Promedio Mensual",
                "Score Financiero",
                "% Gastos / Ingresos",
                "% Ahorro"
            ],
            "Valor": [
                ingresos,
                gastos,
                balance,
                promedio_mensual,
                score,
                abs(gastos) / ingresos if ingresos != 0 else 0,
                balance / ingresos if ingresos != 0 else 0
            ]
        })

        dashboard.to_excel(writer, sheet_name="Dashboard", index=False)
        # -------------------------------------------------
        # Aplicar formato a todas las hojas existentes
        # -------------------------------------------------
        for sheet_name in writer.book.sheetnames:
            ws = writer.book[sheet_name]
            formatear_hoja(ws, estilos)

        # -------------------------------------------------
        # Ajustes especiales hoja Resumen
        # -------------------------------------------------
        ws_res = writer.book['Resumen']

        # Número de movimientos (fila 5) y Score (fila 8)
        ws_res.cell(row=5, column=2).number_format = '0'
        ws_res.cell(row=5, column=2).alignment = estilos['align_center']  # fila de movimientos
        ws_res.cell(row=8, column=2).number_format = '0'
        ws_res.cell(row=8, column=2).alignment = estilos['align_center']  # fila de score

        # -------------------------------------------------
        # Formato condicional Resumen_Mensual
        # -------------------------------------------------
        if 'Resumen_Mensual' in writer.book.sheetnames:
            ws_rm = writer.book['Resumen_Mensual']
            aplicar_formato_condicional_resumen_mensual(ws_rm)

        # -------------------------------------------------
        # Formato especial Dashboard
        # -------------------------------------------------
        if "Dashboard" in writer.book.sheetnames:
            ws_dash = writer.book["Dashboard"]

            # Filas con porcentajes (6 y 7)
            for row in [7, 8]:
                ws_dash.cell(row=row, column=2).number_format = "0.00%"

            # Score financiero (fila 6)
            ws_dash.cell(row=6, column=2).number_format = "0"

    return str(output_path)
