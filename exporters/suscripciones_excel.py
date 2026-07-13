from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from domain.suscripciones.constants import DIAS_ALERTA


def _aplicar_estilo_excel(workbook: Any, df: pd.DataFrame) -> None:
    """Aplica un estilo profesional y sobrio a las hojas exportadas."""
    header_fill = PatternFill(fill_type="solid", fgColor="E8EEF5")
    alternate_fill = PatternFill(fill_type="solid", fgColor="F7FAFC")
    header_font = Font(name="Calibri", size=10, bold=True, color="23364D")
    body_font = Font(name="Calibri", size=10, color="23364D")
    border = Border(
        left=Side(style="thin", color="D0D7DE"),
        right=Side(style="thin", color="D0D7DE"),
        top=Side(style="thin", color="D0D7DE"),
        bottom=Side(style="thin", color="D0D7DE"),
    )
    align_center = Alignment(horizontal="center", vertical="center")

    colores = {
        "Vencida": PatternFill(fill_type="solid", fgColor="FDE2E2"),
        "Vence pronto": PatternFill(fill_type="solid", fgColor="FDE2E2"),
        "Vigente": PatternFill(fill_type="solid", fgColor="EAF7EA"),
    }

    for ws in workbook.worksheets:
        ws.sheet_view.zoomScale = 90
        ws.auto_filter.ref = ws.dimensions

        if ws.title == "Detalle" and "Dias_Vencimiento" in df.columns:
            for row_idx, row in enumerate(
                ws.iter_rows(min_row=2, max_row=ws.max_row), start=2
            ):
                dias = None
                for cell in row:
                    if cell.column == 1:
                        continue
                if ws.cell(row=row_idx, column=1).value is not None:
                    dias_cell = ws.cell(row=row_idx, column=1)
                    dias = dias_cell.value
                if dias is None:
                    continue
                try:
                    dias_num = float(dias)
                except (TypeError, ValueError):
                    continue

                if dias_num < 0:
                    estado = "Vencida"
                elif dias_num <= DIAS_ALERTA:
                    estado = "Vence pronto"
                else:
                    estado = "Vigente"

                for cell in row:
                    cell.fill = colores[estado]

        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = align_center
                else:
                    cell.font = body_font
                    if cell.row % 2 == 0:
                        cell.fill = alternate_fill

        for col_idx, col in enumerate(ws.columns, start=1):
            values = [cell.value for cell in col if cell.value is not None]
            max_length = max((len(str(v)) for v in values), default=10)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(
                max(12, max_length + 2), 50
            )


def exportar_suscripciones_excel(
    df: pd.DataFrame, output_path: str | Path | None = None
) -> str:
    """Exporta un reporte de suscripciones a Excel con formato profesional."""
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        resumen = pd.DataFrame(
            {
                "Indicador": [
                    "Total de suscripciones",
                    "Costo mensual total",
                    "Costo anual total",
                    "Vence pronto",
                    "Vencidas",
                ],
                "Valor": [
                    len(df),
                    df["Costo_Mensual"].sum() if "Costo_Mensual" in df.columns else 0,
                    df["Costo_Anual"].sum() if "Costo_Anual" in df.columns else 0,
                    df["Dias_Vencimiento"].le(DIAS_ALERTA).sum()
                    if "Dias_Vencimiento" in df.columns
                    else 0,
                    df["Estado_Vencimiento"].eq("Vencida").sum()
                    if "Estado_Vencimiento" in df.columns
                    else 0,
                ],
            }
        )

        resumen.to_excel(writer, sheet_name="Resumen", index=False)
        df.to_excel(writer, sheet_name="Detalle", index=False)

        if "Proveedor" in df.columns:
            agrupado = (
                df.groupby("Proveedor", as_index=False)["Costo_Mensual"]
                .sum()
                .sort_values("Costo_Mensual", ascending=False)
            )
            agrupado.to_excel(writer, sheet_name="Por_Proveedor", index=False)

        if "Grupo" in df.columns:
            por_grupo = (
                df.groupby("Grupo", as_index=False)["Costo_Mensual"]
                .sum()
                .sort_values("Costo_Mensual", ascending=False)
            )
            por_grupo.to_excel(writer, sheet_name="Por_Grupo", index=False)

        workbook = writer.book
        _aplicar_estilo_excel(workbook, df)

    output.seek(0)

    if output_path is None:
        output_path = Path("dist") / "suscripciones_reporte.xlsx"
    else:
        output_path = Path(output_path)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("wb") as fh:
        fh.write(output.getvalue())

    return str(output_path)
