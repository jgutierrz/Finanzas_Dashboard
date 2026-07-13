from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# =========================================================
# FORMATOS
# =========================================================

FORMATO_MONEDA = "[$S/] #,##0.00"
FORMATO_FECHA = "dd/mm/yyyy"
FORMATO_PORCENTAJE = "0.00%"


# =========================================================
# COLUMNAS ESPECIALES
# =========================================================

COLUMNAS_MONETARIAS = {
    "monto",
    "Ingresos",
    "Gastos",
    "Balance",
    "Promedio mensual",
}

COLUMNAS_PORCENTAJE = {
    "Pct_Gastos",
    "Pct_Ahorro",
    "Variacion_Balance",
}


# =========================================================
# ESTILOS
# =========================================================

FONT_BASE = Font(
    name="Calibri",
    size=10,
    color="23364D",
)

FONT_HEADER = Font(
    name="Calibri",
    size=10,
    color="23364D",
    bold=True,
)


def crear_estilos() -> dict:
    """
    Devuelve un diccionario con los estilos reutilizables.
    """

    fill_header = PatternFill(
        fill_type="solid",
        fgColor="DCEAF7",
    )

    fill_alternating = PatternFill(
        fill_type="solid",
        fgColor="F6FAFD",
    )

    align_center = Alignment(
        horizontal="center",
        vertical="center",
    )

    align_right = Alignment(
        horizontal="right",
        vertical="center",
    )

    border_thin = Border(
        left=Side(style="thin", color="D0D7DE"),
        right=Side(style="thin", color="D0D7DE"),
        top=Side(style="thin", color="D0D7DE"),
        bottom=Side(style="thin", color="D0D7DE"),
    )

    return {
        "fill_header": fill_header,
        "fill_alternating": fill_alternating,
        "font_header": FONT_HEADER,
        "align_center": align_center,
        "align_right": align_right,
        "border_thin": border_thin,
    }


# =========================================================
# FORMATO GENERAL DE HOJAS
# =========================================================


def formatear_hoja(
    ws: Worksheet,
    estilos: dict,
    header_row: int = 1,
) -> None:
    """
    Aplica formato estándar a una hoja Excel.
    """

    fill_header = estilos["fill_header"]
    fill_alternating = estilos["fill_alternating"]
    font_header = estilos["font_header"]
    align_center = estilos["align_center"]
    align_right = estilos["align_right"]
    border_thin = estilos["border_thin"]

    # -----------------------------------------------------
    # Configuración general
    # -----------------------------------------------------

    ws.sheet_view.zoomScale = 90

    if header_row > 0:
        ws.freeze_panes = f"A{header_row + 1}"

    ws.auto_filter.ref = ws.dimensions

    # -----------------------------------------------------
    # Fuente y bordes
    # -----------------------------------------------------

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border_thin

            if cell.row != header_row:
                cell.font = FONT_BASE
                if cell.row % 2 == 0:
                    cell.fill = fill_alternating

    # -----------------------------------------------------
    # Encabezado
    # -----------------------------------------------------

    for cell in ws[header_row]:
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_center
        cell.border = border_thin

    # -----------------------------------------------------
    # Formatos y ancho automático
    # -----------------------------------------------------

    for col_idx, col in enumerate(ws.columns, start=1):
        header = ws.cell(row=header_row, column=col_idx).value

        max_length = 0

        for cell in col:
            if cell.row <= header_row:
                continue

            # ---------------------------------------------
            # Moneda
            # ---------------------------------------------

            if isinstance(cell.value, (int, float)) and header in COLUMNAS_MONETARIAS:
                cell.number_format = FORMATO_MONEDA
                cell.alignment = align_right

            # ---------------------------------------------
            # Porcentaje
            # ---------------------------------------------

            if isinstance(cell.value, (int, float)) and header in COLUMNAS_PORCENTAJE:
                cell.number_format = FORMATO_PORCENTAJE
                cell.alignment = align_right

            # ---------------------------------------------
            # Fecha
            # ---------------------------------------------

            if header == "fecha" and cell.value is not None:
                cell.number_format = FORMATO_FECHA

            # ---------------------------------------------
            # Ancho automático
            # ---------------------------------------------

            if cell.value is not None:
                longitud = len(str(cell.value))
                max_length = max(max_length, longitud)

        col_letter = get_column_letter(col_idx)

        ws.column_dimensions[col_letter].width = max(12, min(max_length + 2, 50))


# =========================================================
# FORMATO CONDICIONAL
# =========================================================


def aplicar_formato_condicional_resumen_mensual(
    ws: Worksheet,
    header_row: int = 1,
) -> None:
    """
    Formato condicional para la hoja Resumen_Mensual.
    """

    col_map = {cell.value: cell.column for cell in ws[header_row]}

    fill_rojo = PatternFill(
        fill_type="solid",
        fgColor="FDE2E2",
    )

    # -----------------------------------------------------
    # Balance negativo
    # -----------------------------------------------------

    if "Balance" in col_map:
        col = get_column_letter(col_map["Balance"])

        ws.conditional_formatting.add(
            f"{col}{header_row + 1}:{col}{ws.max_row}",
            CellIsRule(
                operator="lessThan",
                formula=["0"],
                fill=fill_rojo,
            ),
        )

    # -----------------------------------------------------
    # % Gastos > 80%
    # -----------------------------------------------------

    if "Pct_Gastos" in col_map:
        col = get_column_letter(col_map["Pct_Gastos"])

        ws.conditional_formatting.add(
            f"{col}{header_row + 1}:{col}{ws.max_row}",
            CellIsRule(
                operator="greaterThan",
                formula=["0.8"],
                fill=fill_rojo,
            ),
        )

    # -----------------------------------------------------
    # % Ahorro < 0
    # -----------------------------------------------------

    if "Pct_Ahorro" in col_map:
        col = get_column_letter(col_map["Pct_Ahorro"])

        ws.conditional_formatting.add(
            f"{col}{header_row + 1}:{col}{ws.max_row}",
            CellIsRule(
                operator="lessThan",
                formula=["0"],
                fill=fill_rojo,
            ),
        )
